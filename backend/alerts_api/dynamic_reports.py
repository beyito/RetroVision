import os
import json
import logging
import requests
from datetime import datetime
from io import BytesIO
from collections import defaultdict

from django.utils.dateparse import parse_datetime
from django.utils import timezone
from datetime import timedelta

from .models import Telemetria_Afluencia, SecurityAlert

logger = logging.getLogger("RetroVision.DynamicReports")

def call_gemini(prompt: str, model: str) -> str:
    api_key = os.getenv("AI_API_KEY")
    if not api_key:
        raise ValueError("La variable de entorno AI_API_KEY no está configurada en el servidor.")
        
    url = f"https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent?key={api_key}"
    headers = {"Content-Type": "application/json"}
    payload = {
        "contents": [{
            "parts": [{
                "text": prompt
            }]
        }],
        "generationConfig": {
            "responseMimeType": "application/json"
        }
    }
    
    try:
        response = requests.post(url, json=payload, headers=headers, timeout=45)
        response.raise_for_status()
        resp_json = response.json()
        
        # Extract text
        text_response = resp_json['candidates'][0]['content']['parts'][0]['text']
        return text_response
    except requests.exceptions.RequestException as e:
        logger.error(f"Gemini API request failed: {e}")
        raise ValueError(f"Error al conectar con la API de Gemini: {str(e)}")
    except (KeyError, IndexError) as e:
        logger.error(f"Gemini API response format invalid: {e}")
        raise ValueError("La API de Gemini devolvió una respuesta con formato inválido.")

def clean_and_parse_json(text_response: str) -> dict:
    text_response = text_response.strip()
    if text_response.startswith("```json"):
        text_response = text_response[7:]
    elif text_response.startswith("```"):
        text_response = text_response[3:]
    if text_response.endswith("```"):
        text_response = text_response[:-3]
    text_response = text_response.strip()
    return json.loads(text_response)

def extract_filters_with_gemini(user_prompt: str, current_time_str: str, model: str) -> dict:
    prompt = f"""
    Analiza la siguiente consulta en lenguaje natural sobre reportes de RetroVision (una plataforma de seguridad y analítica de retail).
    La fecha y hora actuales son: {current_time_str} (UTC).
    Consulta del usuario: "{user_prompt}"

    Debes extraer los parámetros de consulta y rango de fechas para buscar en la base de datos.
    Retorna únicamente un objeto JSON con los siguientes campos (sin texto adicional, sin formato markdown):
    {{
      "start_date": "string (formato ISO 8601, ej. '2026-06-17T00:00:00Z')",
      "end_date": "string (formato ISO 8601, ej. '2026-06-24T23:59:59Z')",
      "camera_id": "string o null (si especificó una cámara en particular)",
      "sector": "string o null (si especificó un sector en particular como 'carnes', 'lacteos')",
      "subject": "string ('telemetry', 'alerts' o 'both')"
    }}

    Notas:
    - Si el usuario dice "hoy", el rango es desde el inicio del día actual (00:00:00) hasta ahora.
    - Si dice "ayer", el rango es todo el día anterior completo.
    - Si no especifica rango de fechas, asume los últimos 7 días como valor por defecto.
    - Traduce expresiones como "último mes", "última semana", "últimas 24 horas" a rangos ISO 8601 relativos a {current_time_str}.
    """
    
    try:
        raw_resp = call_gemini(prompt, model)
        return clean_and_parse_json(raw_resp)
    except Exception as e:
        logger.warning(f"Failed to extract filters via Gemini: {e}. Using defaults.")
        now = timezone.now()
        start = now - timedelta(days=7)
        return {
            "start_date": start.isoformat(),
            "end_date": now.isoformat(),
            "camera_id": None,
            "sector": None,
            "subject": "both"
        }

def compile_metrics(start_date_str, end_date_str, camera_ids, filter_camera_id=None, filter_sector=None):
    start_date = parse_datetime(start_date_str)
    end_date = parse_datetime(end_date_str)
    
    if not start_date:
        start_date = timezone.now() - timedelta(days=7)
    if not end_date:
        end_date = timezone.now()

    # Base filters
    telemetry_qs = Telemetria_Afluencia.objects.filter(
        timestamp__gte=start_date,
        timestamp__lte=end_date,
        camera_id__in=camera_ids
    )
    
    alerts_qs = SecurityAlert.objects.filter(
        timestamp__gte=start_date,
        timestamp__lte=end_date,
        camera_id__in=camera_ids
    )
    
    if filter_camera_id:
        telemetry_qs = telemetry_qs.filter(camera_id=filter_camera_id)
        alerts_qs = alerts_qs.filter(camera_id=filter_camera_id)
        
    record_list = list(telemetry_qs.order_by('timestamp'))
    alerts_list = list(alerts_qs.order_by('-timestamp'))
    
    total_records = len(record_list)
    
    # 1. Telemetry Aggregation
    telemetry_summary = {}
    if total_records > 0:
        total_queue_people = 0
        total_wait_time = 0.0
        max_wait_time = 0.0
        saturated_records_count = 0
        
        sector_totals = defaultdict(int)
        sector_maxes = defaultdict(int)
        sector_record_counts = defaultdict(int)
        
        hourly_inflow_by_hour = defaultdict(int)
        daily_inflow_by_day = defaultdict(int)
        
        prev_entrantes_by_cam = {}
        
        for r in record_list:
            cam_id = r.camera_id
            
            # Queue metrics
            total_queue_people += r.personas_en_cola
            total_wait_time += r.tiempo_espera_promedio
            if r.tiempo_espera_promedio > max_wait_time:
                max_wait_time = r.tiempo_espera_promedio
            if r.alerta_cola_activa or r.personas_en_cola >= 3:
                saturated_records_count += 1
                
            # Sectores metrics
            sectores = r.sectores or {}
            if isinstance(sectores, dict):
                for name, count in sectores.items():
                    normalized_name = str(name).strip().capitalize() if name else "Desconocido"
                    if filter_sector and normalized_name.lower() != filter_sector.lower():
                        continue
                    val = int(count or 0)
                    sector_totals[normalized_name] += val
                    sector_record_counts[normalized_name] += 1
                    if val > sector_maxes[normalized_name]:
                        sector_maxes[normalized_name] = val
                        
            # Flow differences
            val_entrantes = r.personas_entrantes
            hour = r.timestamp.hour
            day_name = r.timestamp.strftime("%A")
            
            if cam_id in prev_entrantes_by_cam:
                diff = val_entrantes - prev_entrantes_by_cam[cam_id]
                if diff > 0:
                    hourly_inflow_by_hour[hour] += diff
                    daily_inflow_by_day[day_name] += diff
            prev_entrantes_by_cam[cam_id] = val_entrantes
            
        avg_people_in_queue = total_queue_people / total_records
        avg_wait_time = total_wait_time / total_records
        saturation_percentage = (saturated_records_count / total_records) * 100.0
        
        sectors_metrics = {}
        for name in sector_totals.keys():
            sectors_metrics[name] = {
                "avg_occupancy": round(sector_totals[name] / sector_record_counts[name], 2),
                "max_occupancy": sector_maxes[name]
            }
            
        # Daily mapping
        day_mapping = {
            "Monday": "Lunes", "Tuesday": "Martes", "Wednesday": "Miércoles",
            "Thursday": "Jueves", "Friday": "Viernes", "Saturday": "Sábado", "Sunday": "Domingo"
        }
        
        daily_inflow = {}
        for eng_name, esp_name in day_mapping.items():
            daily_inflow[esp_name] = daily_inflow_by_day.get(eng_name, 0)
            
        telemetry_summary = {
            "total_records_analyzed": total_records,
            "total_visitors_estimated": sum(daily_inflow.values()),
            "avg_people_in_queue": round(avg_people_in_queue, 1),
            "avg_wait_time_seconds": round(avg_wait_time, 1),
            "max_wait_time_seconds": round(max_wait_time, 1),
            "saturation_percentage": round(saturation_percentage, 1),
            "sectors": sectors_metrics,
            "daily_inflow": daily_inflow
        }
    else:
        telemetry_summary = {
            "total_records_analyzed": 0,
            "total_visitors_estimated": 0,
            "avg_people_in_queue": 0.0,
            "avg_wait_time_seconds": 0.0,
            "max_wait_time_seconds": 0.0,
            "saturation_percentage": 0.0,
            "sectors": {},
            "daily_inflow": {}
        }
        
    # 2. Alerts Aggregation
    rule_counts = defaultdict(int)
    recent_alerts = []
    
    for alert in alerts_list[:30]:
        rules = alert.rules_triggered
        rules_list = rules if isinstance(rules, list) else [str(rules)]
        
        if filter_sector and alert.zona.lower() != filter_sector.lower():
            continue
            
        for r in rules_list:
            normalized_rule = str(r).strip()
            if normalized_rule:
                rule_counts[normalized_rule] += 1
                
        recent_alerts.append({
            "timestamp": alert.timestamp.isoformat(),
            "camera_id": alert.camera_id,
            "risk_score": alert.risk_score,
            "rules": rules_list,
            "zona": alert.zona
        })
        
    all_rule_counts = defaultdict(int)
    for alert in alerts_list:
        if filter_sector and alert.zona.lower() != filter_sector.lower():
            continue
        rules = alert.rules_triggered
        rules_list = rules if isinstance(rules, list) else [str(rules)]
        for r in rules_list:
            normalized_rule = str(r).strip()
            if normalized_rule:
                all_rule_counts[normalized_rule] += 1
                
    security_summary = {
        "total_alerts": len(alerts_list) if not filter_sector else len(recent_alerts),
        "rule_breakdown": dict(all_rule_counts),
        "recent_alerts": recent_alerts[:15]
    }
    
    return {
        "telemetry": telemetry_summary,
        "security": security_summary
    }

def generate_narrative_with_gemini(user_prompt: str, start_date_str: str, end_date_str: str, data_summary: dict, model: str) -> dict:
    prompt = f"""
    Actúa como un experto analista de retail y seguridad para la plataforma RetroVision.
    El usuario ha solicitado un reporte dinámico con la siguiente consulta: "{user_prompt}"

    Hemos consultado la base de datos de RetroVision en el rango de {start_date_str} a {end_date_str} y compilado el siguiente conjunto de datos estructurados:
    {json.dumps(data_summary, indent=2, ensure_ascii=False)}

    Por favor, genera un reporte detallado y profesional en base a estos datos.
    Debes responder únicamente con un objeto JSON estructurado con los siguientes campos (sin texto explicativo fuera del JSON, sin prefijos de markdown):
    {{
      "title": "Título profesional del reporte en español",
      "executive_summary": "Resumen ejecutivo detallado (mínimo 3 oraciones) analizando los hallazgos en español.",
      "kpis": [
        {{ "label": "Nombre del KPI (ej. Afluencia Total)", "value": "Valor con unidad (ej. 345 personas o 25s)" }}
      ],
      "table": {{
        "headers": ["Columna 1", "Columna 2", ...],
        "rows": [
          ["Valor 1a", "Valor 1b", ...],
          ["Valor 2a", "Valor 2b", ...]
        ]
      }},
      "recommendations": [
        "Recomendación accionable 1 en base a los datos",
        "Recomendación accionable 2 en base a los datos"
      ],
      "markdown_content": "El reporte completo formateado en Markdown enriquecido con tablas, listas y secciones."
    }}

    Instrucciones para la tabla:
    - Si la consulta trata sobre sectores, la tabla debe mostrar métricas por sector (Afluencia promedio, Afluencia máxima).
    - Si trata sobre afluencia horaria o diaria, la tabla debe mostrar el flujo por horas o días.
    - Si trata sobre alertas de seguridad, la tabla debe listar las alertas recientes o anomalías por regla.
    - La tabla debe tener de 3 a 10 filas de datos relevantes.

    Asegúrate de que el JSON sea válido y todos los textos estén en español.
    """
    
    raw_resp = call_gemini(prompt, model)
    return clean_and_parse_json(raw_resp)

def generate_pdf_report(report_data: dict) -> bytes:
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=letter,
        rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40
    )
    
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'ReportTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=18,
        leading=22,
        textColor=colors.HexColor('#06b6d4'), # Cyan
        spaceAfter=15
    )
    
    h2_style = ParagraphStyle(
        'SectionHeader',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=12,
        leading=16,
        textColor=colors.HexColor('#6366f1'), # Indigo
        spaceBefore=12,
        spaceAfter=6
    )
    
    body_style = ParagraphStyle(
        'ReportBody',
        parent=styles['BodyText'],
        fontName='Helvetica',
        fontSize=9,
        leading=13,
        textColor=colors.HexColor('#374151'), # Slate
        spaceAfter=8
    )
    
    story = []
    
    # Header
    story.append(Paragraph("RETROVISION - INTELIGENCIA ARTIFICIAL", ParagraphStyle('Sub', fontName='Helvetica-Bold', fontSize=8, leading=10, textColor=colors.HexColor('#9ca3af'))))
    story.append(Spacer(1, 4))
    story.append(Paragraph(report_data.get('title', 'Reporte Analítico'), title_style))
    story.append(Spacer(1, 8))
    
    # Executive Summary
    story.append(Paragraph("Resumen Ejecutivo", h2_style))
    story.append(Paragraph(report_data.get('executive_summary', ''), body_style))
    story.append(Spacer(1, 10))
    
    # KPIs
    kpis = report_data.get('kpis', [])
    if kpis:
        story.append(Paragraph("Indicadores Clave", h2_style))
        kpi_data = []
        kpi_row = []
        for k in kpis:
            kpi_row.append(Paragraph(f"<b>{k.get('label', '')}</b><br/>{k.get('value', '')}", body_style))
        kpi_data.append(kpi_row)
        
        kpi_table = Table(kpi_data, colWidths=[530/len(kpis)]*len(kpis))
        kpi_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#f3f4f6')),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor('#e5e7eb')),
            ('BOX', (0,0), (-1,-1), 1, colors.HexColor('#d1d5db')),
            ('TOPPADDING', (0,0), (-1,-1), 8),
            ('BOTTOMPADDING', (0,0), (-1,-1), 8),
        ]))
        story.append(kpi_table)
    story.append(Spacer(1, 12))
    
    # Table
    table_dict = report_data.get('table', {})
    headers = table_dict.get('headers', [])
    rows = table_dict.get('rows', [])
    if headers and rows:
        story.append(Paragraph("Datos Detallados", h2_style))
        table_data = []
        
        header_row = [Paragraph(f"<b>{h}</b>", ParagraphStyle('HeaderStyle', parent=body_style, fontName='Helvetica-Bold', textColor=colors.white)) for h in headers]
        table_data.append(header_row)
        
        for r in rows:
            table_data.append([Paragraph(str(cell), body_style) for cell in r])
            
        col_width = 530 / len(headers)
        data_table = Table(table_data, colWidths=[col_width]*len(headers))
        
        t_style = [
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1e293b')),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
            ('TOPPADDING', (0,0), (-1,-1), 6),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#cbd5e1')),
        ]
        
        for i in range(1, len(table_data)):
            if i % 2 == 0:
                t_style.append(('BACKGROUND', (0,i), (-1,i), colors.HexColor('#f8fafc')))
            else:
                t_style.append(('BACKGROUND', (0,i), (-1,i), colors.white))
                
        data_table.setStyle(TableStyle(t_style))
        story.append(data_table)
    story.append(Spacer(1, 12))
    
    # Recommendations
    recs = report_data.get('recommendations', [])
    if recs:
        story.append(Paragraph("Recomendaciones Estratégicas", h2_style))
        for r in recs:
            story.append(Paragraph(f"• {r}", body_style))
            
    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()

def generate_excel_report(report_data: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte IA"
    
    # Show Gridlines
    ws.views.sheetView[0].showGridLines = True
    
    # Title
    ws["A1"] = report_data.get("title", "Reporte de Inteligencia").upper()
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="06b6d4")
    
    # Executive Summary
    ws["A3"] = "RESUMEN EJECUTIVO"
    ws["A3"].font = Font(name="Calibri", size=11, bold=True, color="6366f1")
    ws["A4"] = report_data.get("executive_summary", "")
    ws["A4"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A4:F4")
    ws.row_dimensions[4].height = 45
    
    # KPIs
    ws["A6"] = "INDICADORES CLAVE"
    ws["A6"].font = Font(name="Calibri", size=11, bold=True, color="6366f1")
    
    kpis = report_data.get("kpis", [])
    row_idx = 7
    for k in kpis:
        ws.cell(row=row_idx, column=1, value=k.get("label", "")).font = Font(bold=True)
        ws.cell(row=row_idx, column=2, value=k.get("value", ""))
        row_idx += 1
        
    row_idx += 1 # Space
    
    # Table
    table_dict = report_data.get("table", {})
    headers = table_dict.get("headers", [])
    rows = table_dict.get("rows", [])
    
    if headers and rows:
        ws.cell(row=row_idx, column=1, value="DATOS DETALLADOS").font = Font(name="Calibri", size=11, bold=True, color="6366f1")
        row_idx += 1
        
        # Headers
        col_idx = 1
        header_fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="ffffff")
        for h in headers:
            cell = ws.cell(row=row_idx, column=col_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            col_idx += 1
            
        row_idx += 1
        
        # Rows
        border_thin = Border(
            left=Side(style='thin', color='cbd5e1'),
            right=Side(style='thin', color='cbd5e1'),
            top=Side(style='thin', color='cbd5e1'),
            bottom=Side(style='thin', color='cbd5e1')
        )
        
        for r in rows:
            col_idx = 1
            for val in r:
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = border_thin
                col_idx += 1
            row_idx += 1
            
    row_idx += 1 # Space
    
    # Recommendations
    recs = report_data.get("recommendations", [])
    if recs:
        ws.cell(row=row_idx, column=1, value="RECOMENDACIONES Y SUGERENCIAS").font = Font(name="Calibri", size=11, bold=True, color="6366f1")
        row_idx += 1
        for r in recs:
            ws.cell(row=row_idx, column=1, value=f"• {r}")
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=6)
            ws.cell(row=row_idx, column=1).alignment = Alignment(wrap_text=True)
            row_idx += 1
            
    # Auto column widths
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.coordinate in ["A4"] or (cell.row >= row_idx - len(recs) and len(recs) > 0):
                continue
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

def generate_dynamic_report(user, camera_ids, user_prompt: str, format_type: str, model_name: str) -> tuple:
    current_time_str = datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")
    filters = extract_filters_with_gemini(user_prompt, current_time_str, model_name)
    
    start_date = filters.get("start_date")
    end_date = filters.get("end_date")
    filter_camera_id = filters.get("camera_id")
    filter_sector = filters.get("sector")
    
    if filter_camera_id and filter_camera_id not in camera_ids:
        filter_camera_id = None
        
    data_summary = compile_metrics(
        start_date, end_date, camera_ids, 
        filter_camera_id=filter_camera_id, 
        filter_sector=filter_sector
    )
    
    data_summary["metadata"] = {
        "user_query": user_prompt,
        "extracted_filters": filters,
        "current_time_utc": current_time_str
    }
    
    report_data = generate_narrative_with_gemini(user_prompt, start_date, end_date, data_summary, model_name)
    
    if format_type == "pdf":
        file_data = generate_pdf_report(report_data)
        mime_type = "application/pdf"
        file_ext = "pdf"
        return file_data, mime_type, file_ext
    elif format_type == "excel":
        file_data = generate_excel_report(report_data)
        mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        file_ext = "xlsx"
        return file_data, mime_type, file_ext
    else:
        return report_data, "application/json", "json"
